# robot_logic.py
from PyQt5.QtWidgets import QTableWidgetItem, QMessageBox, QDialog, QInputDialog, QLineEdit
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QColor
from openpyxl import Workbook
from robot_dialog import RobotDialog
from history_dialog import HistoryDialog
from datetime import datetime, timedelta

class RobotLogic:
    def __init__(self, ui, service):
        """
        ui: экземпляр окна/виджета (например, RobotTable из ui.py)
        service: экземпляр RobotService (services.RobotService)
        """
        self.ui = ui
        self.service = service

        self.headers = [
            "Модель", "Серийный № робота", "Серийный № контроллера","Дата поступления", "Текущий статус",
            "Описание неисправности", "Причина поломки", "Проведенные работы",
            "Планируемые работы", "Необходимые запчасти", "Примечания"
        ]

        self.db_fields = [
            "model", "robot_sn", "controller_sn","arrival_date", "status",
            "fault_description", "fault_reason", "tasks_done",
            "tasks_required", "required_parts", "notes"
        ]

        self.start_auto_refresh()

    def start_auto_refresh(self):
        """Автообновление таблицы каждую минуту."""
        self.timer = QTimer()
        self.timer.timeout.connect(self.load_data)
        self.timer.start(60000)  # 60 секунд

    def show_history(self):
        """Открывает диалог истории."""
        dialog = HistoryDialog(service=self.service, parent=self.ui)
        dialog.exec_()

    def clear_history(self):
        """Очистка всей истории через сервис с подтверждением пароля."""
        from PyQt5.QtWidgets import QInputDialog, QLineEdit

        password, ok = QInputDialog.getText(
            self.ui, "Пароль", "Введите пароль для очистки истории:", QLineEdit.Password
        )
        if not ok:
            return

        expected = self.service.get_history_clear_password()  # или self.service.config.get("history_clear_password")
        if expected is None:
            QMessageBox.critical(self.ui, "Ошибка", "Пароль не задан в config.json")
            return

        if password != expected:
            QMessageBox.warning(self.ui, "Ошибка", "Неверный пароль!")
            return

        reply = QMessageBox.question(
            self.ui, "Подтверждение",
            "Вы уверены, что хотите очистить всю историю?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.service.clear_history()
            self.load_data()  # или self.show_history()
            QMessageBox.information(self.ui, "История", "🧹 История успешно очищена.")


    def create_table_item(self, value, field=None):
        """Создаёт QTableWidgetItem с правильными флагами и раскраской по статусу.
        Добавляет конвертацию даты к нужному формату"""
        if field == "arrival_date" and value:
                if isinstance(value, str):
                    # если из базы пришла строка 'YYYY-MM-DD'
                    parts = value.split("-")
                    if len(parts) == 3:
                        value = f"{parts[2]}.{parts[1]}.{parts[0]}"  # ДД.ММ.ГГГГ
                elif hasattr(value, "strftime"):
                    # если пришёл объект datetime.date или datetime.datetime
                    value = value.strftime("%d.%m.%Y")  # ДД.ММ.ГГГГ

        item = QTableWidgetItem(str(value))
        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
        if field == "status":
            color_map = {
                "Необходим ремонт": QColor("#ffcccc"),
                "Откалиброван": QColor("#8FBC8F"),
                "Тестируется": QColor("#ffffcc"),
                "Протестирован": QColor("#ccffff"),
                "Упакован": QColor("#87CEEB"),
                "Отгружен": QColor("#d0d0d0"),
                "Простаивает": QColor("#DCDCDC")
            }
            color = color_map.get(value, QColor("#ffffff"))
            item.setBackground(color)

        return item

    def load_data(self):
        """Загружает данные из БД и заполняет таблицу UI."""
        try:
            self.ui.table.blockSignals(True)
        except Exception:
            pass

        robots = self.service.get_all_robots() or []
        robots.sort(key=lambda x: x.get('id', 0))

        self.ui.table.setColumnCount(len(self.headers))
        self.ui.table.setHorizontalHeaderLabels(self.headers)
        self.ui.table.setRowCount(len(robots))

        for row_idx, robot in enumerate(robots):
            for col_idx, field in enumerate(self.db_fields):
                value = robot.get(field, "")
                item = self.create_table_item(value, field)
                self.ui.table.setItem(row_idx, col_idx, item)

        try:
            self.ui.table.resizeColumnsToContents()
            self.ui.table.resizeRowsToContents()
        except Exception:
            pass

        try:
            self.ui.table.blockSignals(False)
        except Exception:
            pass

        # Настройка ширины колонки Статус и затемнение строк "Отгружен"
        if "status" in self.db_fields:
            status_column_index = self.db_fields.index("status")
            try:
                self.ui.table.setColumnWidth(status_column_index, 170)
            except Exception:
                pass

            for row_idx, robot in enumerate(robots):
                if robot.get("status") == "Отгружен":
                    for col_idx in range(self.ui.table.columnCount()):
                        cell = self.ui.table.item(row_idx, col_idx)
                        if cell:
                            cell.setBackground(QColor(200, 200, 200))

        self.apply_filters()

    def apply_filters(self):
        """Фильтрация по поиску + скрытие отгруженных"""
        query = self.ui.search_input.text().lower()
        hide_shipped = self.ui.hide_shipped_checkbox.isChecked()

        status_col = self.db_fields.index("status")

        for row in range(self.ui.table.rowCount()):
            visible = True

            # 🔍 Поиск
            if query:
                visible = False
                for col in range(self.ui.table.columnCount()):
                    item = self.ui.table.item(row, col)
                    if item and query in item.text().lower():
                        visible = True
                        break

            # 🚚 Скрытие отгруженных
            if hide_shipped:
                status_item = self.ui.table.item(row, status_col)
                if status_item and status_item.text() == "Отгружен":
                    visible = False

            self.ui.table.setRowHidden(row, not visible)

    def add_robot(self):
        """Открывает диалог добавления робота и добавляет через сервис."""
        dialog = RobotDialog(parent=self.ui)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            self.service.add_robot(data)
            self.load_data()

    def edit_robot(self):
        """Редактирование выбранной строки через диалог."""
        selected_row = self.ui.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self.ui, "Нет выбора", "Выберите строку для редактирования")
            return

        robots = self.service.get_all_robots() or []
        if selected_row >= len(robots):
            QMessageBox.warning(self.ui, "Ошибка", "Выбранная строка вне диапазона")
            return

        robot = robots[selected_row]
        dialog = RobotDialog(robot_data=robot, parent=self.ui)
        if dialog.exec_() == QDialog.Accepted:
            updated_data = dialog.get_data()
            robot_id = robot.get('id')
            for field, value in updated_data.items():
                old_value = robot.get(field)
                if str(old_value or "") != str(value or ""):
                    self.service.update_robot(robot_id, field, value, old_value)
            self.load_data()
            QMessageBox.information(self.ui, "Готово", "✅ Робот обновлён.")

    def delete_robot(self):
        """Удаляет выбранного робота после подтверждения."""
        selected_row = self.ui.table.currentRow()
        if selected_row < 0:
            return

        robots = self.service.get_all_robots() or []
        if selected_row >= len(robots):
            return

        robot_id = robots[selected_row].get('id')
        reply = QMessageBox.question(
            self.ui, "Подтверждение удаления", f"Удалить робота с ID {robot_id}?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.service.delete_robot(robot_id)
            self.load_data()

    def save_changes(self):
        """Подтверждение сохранения всех изменений (для будущей логики редактирования на месте)."""
        reply = QMessageBox.question(self.ui, "Подтверждение", "Сохранить все изменения?", QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        QMessageBox.information(self.ui, "Сохранено", "Изменения сохранены (если были).")

    def export_to_excel(self):
        """Экспорт таблицы роботов в Excel с форматированием и выбором пути сохранения."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from PyQt5.QtWidgets import QFileDialog, QMessageBox
        from datetime import datetime

        # Получаем данные
        robots = self.service.get_all_robots() or []
        if not robots:
            QMessageBox.information(self.ui, "Экспорт", "⚠️ Нет данных для экспорта.")
            return

        # Диалог выбора пути сохранения
        default_name = f"robots_ОТК_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self.ui,
            "Сохранить как",
            default_name,
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return  # пользователь отменил

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Роботы ОТК"

            # === Заголовки ===
            ws.append(self.headers)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            align_center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for col_num, header in enumerate(self.headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border

            # === Данные ===
            status_colors = {
                "Необходим ремонт": "FFFFC7CE",
                "Откалиброван": "FFC6EFCE",
                "Тестируется": "FFFFF2CC",
                "Протестирован": "FFCCFFFF",
                "Упакован": "FF87CEEB",
                "Отгружен": "FFD9D9D9",
                "Простаивает": "FFE7E6E6"
            }

            for row_idx, robot in enumerate(robots, start=2):
                for col_idx, field in enumerate(self.db_fields, start=1):
                    value = robot.get(field, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if field == "status":
                        color = status_colors.get(value, "FFFFFFFF")
                        cell.fill = PatternFill("solid", fgColor=color)

            # === Автоширина ===
            for col_cells in ws.columns:
                col_cells = [c for c in col_cells if c.value is not None]
                if not col_cells:
                    continue
                max_len = max(len(str(c.value)) for c in col_cells)
                col_letter = get_column_letter(col_cells[0].column)
                ws.column_dimensions[col_letter].width = max_len + 2

            # === Заморозка и автофильтр ===
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

            # === Добавляем дату экспорта ===
            ws["A{}".format(ws.max_row + 2)] = f"Экспортировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

            # === Сохранение ===
            wb.save(file_path)
            wb.close()

            QMessageBox.information(self.ui, "Экспорт", f"✅ Данные успешно экспортированы в:\n{file_path}")

        except PermissionError:
            QMessageBox.warning(
                self.ui,
                "Ошибка доступа",
                "⚠️ Невозможно сохранить файл — возможно, он уже открыт в Excel.\n"
                "Закройте файл и попробуйте снова."
            )
        except Exception as e:
            QMessageBox.critical(
                self.ui,
                "Ошибка экспорта",
                f"❌ Не удалось сохранить файл:\n{e}"
            )

